"""
ECサイト(I-PREMIUMPLAZA)から商品情報をスクレイピングするプログラム

2019/10/25
制作 katada
"""
# -*- coding: utf-8 -*-

import bs4
import requests
import openpyxl
import datetime
import shutil

hin_data_all = []#全てのデータ格納先
today = datetime.date.today()

def datasclyping(syouhincode):
    URL = "https://www.premium-plaza.jp/ShouhinKensaku/Shousai/"+syouhincode#引数で商品コードを受け取る
    res = requests.get(URL)
    try:
        res.raise_for_status()
    except Exception as exc:
        print("問題あり:{}".format(exc))

    soup = bs4.BeautifulSoup(res.content, 'lxml')

    span = soup.find_all("span")#spanタグをリストに格納
    td = soup.find_all("td")#tdタグをリストに格納
    ptag = soup.find_all("p")#aタグをリストに格納

    hin_data = []#1件分格納
    try:
        hin_data.append(span[10].text.replace(u"\n",u"").replace(u"\xa0",u"").replace(u"\r",u"").replace(u" ",u""))
        hin_data.append(ptag[0].text.replace(u"\n",u"").replace(u"\xa0",u"").replace(u"\r",u"").replace(u" ",u""))
        hin_data.append(ptag[1].text.replace(u"\n",u"").replace(u"\xa0",u"").replace(u"\r",u"").replace(u" ",u""))
        x = 90
        for data in td[90:114]:
            hin_data.append(td[x].text.replace(u"\n",u"").replace(u"\xa0",u"").replace(u"\r",u"").replace(u" ",u""))
            x+=1

        hin_data_all.append(hin_data)

    except IndexError:
        print("指定した商品コードがうまく読み込めませんでした")
        print(syouhincode)
    




loadwb = openpyxl.load_workbook('hinlist.xlsx') 
loadws = loadwb.worksheets[0]#検索した商品コードをリストに読み込む
loadcoad = []#商品コード格納リスト
maxrow=loadws.max_row

while maxrow >0:
    #エクセルに入力されている商品コードの数だけリストにコードを追加する
    loadcoad.append(str(loadws["A"+str(maxrow)].value))
    maxrow-=1

loadcoad.reverse()#逆順でリストに入ってしまっているので並び替える

for x in range(len(loadcoad)):
    datasclyping(loadcoad[x])

wb = openpyxl.Workbook() 
ws = wb.worksheets[0]
ws.title = '出力データ'+str(today)+'.xlsx'
    
print("出力開始")

for x in range(len(loadcoad)):
    rownum = x + 1
    print(str(x)+"件目")
    ws.cell(row=rownum,column=1).value = loadcoad[x]
    #商品コードをA列に入れる
    for i in range(len(hin_data_all[0])):
        colnum = i + 2
        try:
            ws.cell(row=rownum,column=colnum).value = hin_data_all[x][i]
        except Exception as exc:
            pass

wb.save('出力データ'+str(today)+'.xlsx')

shutil.move('出力データ'+str(today)+'.xlsx',"C:\\Users\seisaku\Desktop\獲得商品情報")

print("出力完了")